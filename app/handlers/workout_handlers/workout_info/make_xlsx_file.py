import datetime
from io import BytesIO

from aiogram.types import file
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from app.database.models import Workout

from app.database.queries import ORMWorkout


async def create_file_with_all_workouts_info() -> str:
    wb = load_workbook(filename='app/handlers/workout_handlers/workout_info/template.xlsx')
    sheet = wb['Sheet1']
    column_range = [2, 3]
    row_range = range(1, 13)
    all_workours_info = await ORMWorkout.get_all_workouts()
    for w in all_workours_info:
        parse_workout_to_excel_table(w=w[0], sheet=sheet, column_range=column_range, row_range=row_range)
        column_range = list(map(lambda x: x+2, column_range))
    filename = f'app/handlers/workout_handlers/workout_info/xlsx_files/{datetime.datetime.now()}.xlsx'
    wb.save(filename=filename)

    return filename


def parse_workout_to_excel_table(w: Workout, sheet: Worksheet, column_range: list, row_range: range) -> None:
    # Записываем id тренировки
    sheet.cell(column=column_range[0], row=row_range[0]).value = w.id

    # Записываем названия столбцов
    sheet.cell(column=column_range[0], row=row_range[1]).value = 'План'
    sheet.cell(column=column_range[1], row=row_range[1]).value = 'Факт'

    # Записываем упражнения
    sheet.cell(column=column_range[0], row=row_range[2]).value = w.deadlift_plan
    sheet.cell(column=column_range[1], row=row_range[2]).value = w.deadlift_actually
    sheet.cell(column=column_range[0], row=row_range[3]).value = w.sqatting_plan
    sheet.cell(column=column_range[1], row=row_range[3]).value = w.sqatting_actually
    sheet.cell(column=column_range[0], row=row_range[4]).value = w.bench_press_plan
    sheet.cell(column=column_range[1], row=row_range[4]).value = w.bench_press_actually
    sheet.cell(column=column_range[0], row=row_range[5]).value = w.standing_barbell_curl_plan
    sheet.cell(column=column_range[1], row=row_range[5]).value = w.standing_barbell_curl_actually
    sheet.cell(column=column_range[0], row=row_range[6]).value = w.pull_up_plan
    sheet.cell(column=column_range[1], row=row_range[6]).value = w.pull_up_actually
    sheet.cell(column=column_range[0], row=row_range[7]).value = w.dumbbell_inclene_bench_press_plan
    sheet.cell(column=column_range[1], row=row_range[7]).value = w.dumbbell_inclene_bench_press_actually
    sheet.cell(column=column_range[0], row=row_range[8]).value = w.military_press_plan
    sheet.cell(column=column_range[1], row=row_range[8]).value = w.military_press_actually
    sheet.cell(column=column_range[0], row=row_range[9]).value = w.lat_pull_down_plan
    sheet.cell(column=column_range[1], row=row_range[9]).value = w.lat_pull_down_actually
    sheet.cell(column=column_range[0], row=row_range[10]).value = w.seated_row_plan
    sheet.cell(column=column_range[1], row=row_range[10]).value = w.seated_row_actually
    sheet.cell(column=column_range[0], row=row_range[11]).value = w.completion
