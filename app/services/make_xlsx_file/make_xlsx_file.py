import asyncio
import datetime

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from app.database.queries import ORMWorkout, ORMPurpose
from app.services.calculate_workout_program.get_exercises_without_purpose import get_exercises_with_purpose


async def create_file_with_all_workouts_info(user_id: int) -> str:
    wb = load_workbook(
        filename='/home/egorshishkin/Desktop/projects/tg_bot_strength_train/app/services/make_xlsx_file/template.xlsx')
    sheet = wb['Sheet1']
    workout_info_about_all_exercises = await get_workout_info_about_all_exercises(user_id)
    for exercise in workout_info_about_all_exercises.items():
        parse_workout_to_excel_table(exercise=exercise[0], w=exercise[1], sheet=sheet)
    filename = (f'/home/egorshishkin/Desktop/projects/tg_bot_strength_train/app/services/make_xlsx_file/xlsx_files'
                f'/{datetime.datetime.now()}.xlsx')
    wb.save(filename=filename)

    return filename


async def get_workout_info_about_all_exercises(user_id: int) -> dict:
    purpose = await ORMPurpose.get_purpose(user_id)
    exercise_with_purpose = get_exercises_with_purpose(purpose)
    all_exercise_info = {}
    for exercise in exercise_with_purpose:
        all_exercise_info[exercise] = await ORMWorkout.get_all_exercise_info(user_id, exercise)

    return all_exercise_info


def parse_workout_to_excel_table(exercise: str, w: list, sheet: Worksheet) -> None:
    start_row = 3
    # Записываем упражнения
    match exercise:
        case 'deadlift':
            for _ in w:
                sheet.cell(column=1, row=start_row).value = _.plan
                sheet.cell(column=2, row=start_row).value = _.fact
                start_row += 1
        case 'squatting':
            for _ in w:
                sheet.cell(column=3, row=start_row).value = _.plan
                sheet.cell(column=4, row=start_row).value = _.fact
                start_row += 1
        case 'bench_press':
            for _ in w:
                sheet.cell(column=5, row=start_row).value = _.plan
                sheet.cell(column=6, row=start_row).value = _.fact
                start_row += 1
        case 'standing_barbell_curl':
            for _ in w:
                sheet.cell(column=7, row=start_row).value = _.plan
                sheet.cell(column=8, row=start_row).value = _.fact
                start_row += 1
        case 'pull_up':
            for _ in w:
                sheet.cell(column=9, row=start_row).value = _.plan
                sheet.cell(column=10, row=start_row).value = _.fact
                start_row += 1
        case 'dumbbell_inclene_bench_press':
            for _ in w:
                sheet.cell(column=11, row=start_row).value = _.plan
                sheet.cell(column=12, row=start_row).value = _.fact
                start_row += 1
        case 'military_press':
            for _ in w:
                sheet.cell(column=13, row=start_row).value = _.plan
                sheet.cell(column=14, row=start_row).value = _.fact
                start_row += 1
        case 'lat_pull_down':
            for _ in w:
                sheet.cell(column=15, row=start_row).value = _.plan
                sheet.cell(column=16, row=start_row).value = _.fact
                start_row += 1
        case 'seated_row':
            for _ in w:
                sheet.cell(column=17, row=start_row).value = _.plan
                sheet.cell(column=18, row=start_row).value = _.fact
                start_row += 1

#
# async def main():
#     # await ORMUser.create_tables()
#     print(await create_file_with_all_workouts_info(1488))
#
#
# if __name__ == "__main__":
#     asyncio.run(main())
